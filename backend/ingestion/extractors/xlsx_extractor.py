from io import BytesIO
import os
import logging
from time import perf_counter

from ingestion.extractors.base import BaseExtractor, ExtractorError, missing_dependency
from ingestion.metadata import (
    build_base_metadata,
    classify_table,
    enrich_metadata,
    normalize_cell_value,
)
from ingestion.schemas import ParsedSection, ParsedTable, ParseResult

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


MAX_TABLE_ROWS = 5000


class XLSXExtractor(BaseExtractor):
    file_type = "xlsx"

    def extract(self, path: str, filename: str) -> ParseResult:
        started_at = perf_counter()
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as exc:
            raise missing_dependency("openpyxl", self.file_type) from exc

        warnings = []
        logger.info("XLSX parse started filename=%s path=%s", filename, path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            path_extension = os.path.splitext(path)[1].lower()
            if not path_extension and isinstance(exc, InvalidFileException):
                logger.warning(
                    "XLSX parse retry without path extension filename=%s path=%s error=%s",
                    filename,
                    path,
                    exc,
                )
                try:
                    with open(path, "rb") as workbook_stream:
                        workbook_bytes = workbook_stream.read()
                    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
                except Exception as retry_exc:
                    raise ExtractorError(f"Failed to read XLSX: {retry_exc}") from retry_exc
            else:
                raise ExtractorError(f"Failed to read XLSX: {exc}") from exc

        document_properties = self._properties(workbook)
        tables = []
        sections = []
        sheet_names = workbook.sheetnames
        logger.info(
            "XLSX workbook opened filename=%s sheets=%s names=%s",
            filename,
            len(sheet_names),
            sheet_names,
        )

        for order, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_started_at = perf_counter()
            logger.info(
                "XLSX sheet start filename=%s sheet=%s index=%s max_row=%s max_column=%s",
                filename,
                worksheet.title,
                order,
                worksheet.max_row,
                worksheet.max_column,
            )
            rows_started_at = perf_counter()
            rows = self._trim_rows([
                [normalize_cell_value(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ])
            logger.info(
                "XLSX sheet rows normalized filename=%s sheet=%s rows=%s in %.2fs",
                filename,
                worksheet.title,
                len(rows),
                perf_counter() - rows_started_at,
            )
            if not rows:
                logger.info("XLSX sheet skipped as empty filename=%s sheet=%s", filename, worksheet.title)
                continue

            headers, data_rows = self._split_header(rows)
            logger.info(
                "XLSX sheet header parsed filename=%s sheet=%s columns=%s data_rows=%s",
                filename,
                worksheet.title,
                len(headers),
                len(data_rows),
            )
            if len(data_rows) > MAX_TABLE_ROWS:
                warnings.append(f"Sheet '{worksheet.title}' was truncated to {MAX_TABLE_ROWS} rows")
                logger.warning(
                    "XLSX sheet truncated filename=%s sheet=%s original_rows=%s max_rows=%s",
                    filename,
                    worksheet.title,
                    len(data_rows),
                    MAX_TABLE_ROWS,
                )
                data_rows = data_rows[:MAX_TABLE_ROWS]

            row_dicts_started_at = perf_counter()
            row_dicts = [self._row_dict(headers, row) for row in data_rows]
            logger.info(
                "XLSX sheet row dicts built filename=%s sheet=%s rows=%s in %.2fs",
                filename,
                worksheet.title,
                len(row_dicts),
                perf_counter() - row_dicts_started_at,
            )
            data_range = f"A1:{get_column_letter(len(headers))}{len(rows)}"
            classification = classify_table(worksheet.title, headers)
            table = ParsedTable(
                name=worksheet.title,
                sheet_name=worksheet.title,
                columns=headers,
                rows=row_dicts,
                data_range=data_range,
                classification=classification,
                meta={"max_row": worksheet.max_row, "max_column": worksheet.max_column},
            )
            tables.append(table)
            sections.append(ParsedSection(
                title=f"Sheet: {worksheet.title}",
                text=(
                    f"Sheet '{worksheet.title}' contains {len(row_dicts)} data rows "
                    f"and {len(headers)} columns."
                ),
                order=order,
                meta={"sheet_name": worksheet.title, "classification": classification},
            ))
            logger.info(
                "XLSX sheet parsed filename=%s sheet=%s classification=%s columns=%s rows=%s in %.2fs",
                filename,
                worksheet.title,
                classification,
                len(headers),
                len(row_dicts),
                perf_counter() - sheet_started_at,
            )

        raw_text = "\n".join(section.text for section in sections)
        metadata = build_base_metadata(
            filename,
            self.file_type,
            document_properties=document_properties,
            sheet_names=sheet_names,
        )
        metadata = enrich_metadata(
            metadata,
            text=raw_text,
            source_type="dataset",
            year_values=(document_properties.get("created"), document_properties.get("modified")),
        )
        try:
            workbook.close()
        except Exception:
            pass
        logger.info(
            "XLSX parse finished filename=%s sheets=%s sections=%s tables=%s warnings=%s raw_text_chars=%s in %.2fs",
            filename,
            len(sheet_names),
            len(sections),
            len(tables),
            len(warnings),
            len(raw_text),
            perf_counter() - started_at,
        )

        return ParseResult(
            filename=filename,
            file_type=self.file_type,
            raw_text=raw_text,
            sections=sections,
            tables=tables,
            metadata=metadata,
            warnings=warnings,
        )

    def _properties(self, workbook):
        props = workbook.properties
        return {
            "title": props.title or None,
            "subject": props.subject or None,
            "creator": props.creator or None,
            "keywords": props.keywords or None,
            "description": props.description or None,
            "category": props.category or None,
            "created": props.created,
            "modified": props.modified,
            "last_modified_by": props.lastModifiedBy or None,
        }

    def _trim_rows(self, rows):
        cleaned = []
        max_width = 0
        for row in rows:
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            if any(value is not None for value in values):
                cleaned.append(values)
                max_width = max(max_width, len(values))
        return [row + [None] * (max_width - len(row)) for row in cleaned]

    def _split_header(self, rows):
        first = rows[0]
        if self._looks_like_header(first):
            return self._headers(first), rows[1:]
        headers = [f"Column {index}" for index in range(1, len(first) + 1)]
        return headers, rows

    def _looks_like_header(self, row):
        values = [value for value in row if value is not None]
        if not values:
            return False
        string_count = sum(isinstance(value, str) for value in values)
        return string_count >= max(1, len(values) // 2)

    def _headers(self, values):
        headers = []
        seen = {}
        for index, value in enumerate(values, start=1):
            header = str(value or f"Column {index}").strip() or f"Column {index}"
            count = seen.get(header.lower(), 0) + 1
            seen[header.lower()] = count
            headers.append(header if count == 1 else f"{header} {count}")
        return headers

    def _row_dict(self, headers, row):
        return {
            header: normalize_cell_value(row[index]) if index < len(row) else None
            for index, header in enumerate(headers)
        }

